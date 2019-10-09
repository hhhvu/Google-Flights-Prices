'''
https://github.com/abhishekchhibber/Gmail-Api-through-Python
'''


from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
import datefinder
import re
import pandas as pd
import numpy as np
import xlrd
import openpyxl
import datetime
from datetime import datetime as dt

import base64
import email
from googleapiclient import errors


args = tools.argparser.parse_args()
args.noauth_local_webserver = True

SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'
store = file.Storage('credentials.json')
creds = store.get()

if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets('C:\\Users\\huong.vu\\Desktop\\Git\\Google-Flights-Prices\\client_secret.json',
                                          SCOPES)
    creds = tools.run_flow(flow, store, args)

service = build('gmail', 'v1', http=creds.authorize(Http()))

# calll the Gmail API, only get 1 of the recent message ids
# First get the message if for the message
results = service.users().messages().list(userId='me',
                                          maxResults=100,  # max record to  obtain
                                          q='from: noreply-travel@google.com label:inbox ').execute()  # include filter for message

time = []
price_now = []
price_before = []
airlines = []

dollar = re.compile(r'\$\d+\,*\d+')
eva = re.compile(r'EVA')

for i in range(results['resultSizeEstimate']):
    # get the message id from the results object
    message_id = results['messages'][i]['id']

    # use the message id to get the actual message, including any attachments
    message = service.users().messages().get(userId='me', id=message_id).execute()
    # print(message['snippet'])
    # for dic in message['payload']['headers']:
    #     if dic['name'] == 'Subject':
    #         print(dic['value'])
    ''' 
    we know the structure of a message variable and the information we need is from ['snippet']
    and ['payload']['headers'], so we can directly go there without a loop.
    Subject and Date are under ['payload']['headers']; however, the object is a list of dictionaries,
    so we cant reference by name. So, we have to use a loop to go through.
    '''

    '''
    If the snippet contains:
        1. 'your tracked flight' with 'EVA Air' => airline = EVA Air
        2. 'your tracked flight' without 'EVA Air' => airline = EVA Air cooperated
        3. 'your tracked flights' => EVA Air and EVA Air cooperated
        4. no 'your track flight(s)' => specify in snippet
    '''
    # prices = [p for p in dollar.finditer(message['snippet'])]
    for dic in message['payload']['headers']:
        if dic['name'] == 'Subject':
            prices = [p.group() for p in dollar.finditer(dic['value'])]
            if prices:
                price_now.append(prices[0])
                price_before.append(prices[1])
            else:
                prices = [p.group() for p in re.finditer(dollar, message['snippet'])]
                price_now.append(prices[0])
                price_before.append(prices[1])
        if dic['name'] == 'Date':
            date = dt.strptime(dic['value'], '%a, %d %b %Y %H:%M:%S %z').strftime('%m/%d/%y')
            time.append(date)

    snippet = message['snippet']
    if re.search('tracked flights', snippet):
        airlines.append('EVA Air & cooeperated')
    elif re.search('tracked flight',snippet) and re.search('EVA Air',snippet):
        airlines.append('EVA Air')
    elif re.search(r'tracked flight[^s]*', snippet):
        airlines.append('EVA cooperated')
    else:
        prices = [p.end() for p in re.finditer(dollar, snippet)]
        spaces = [s.end() for s in re.finditer('\s', snippet[prices[1]:])]
        airlines.append(snippet[prices[1]:(prices[1] + spaces[2])].strip())
        if snippet[prices[1]:(prices[1] + spaces[2])].strip() == 'your tracked':
            print(message_id)



# reorder data
time.reverse()
price_before.reverse()
price_now.reverse()
airlines.reverse()

# bring result into data frame
data = pd.DataFrame(data={'time': time,
                          'price_now': price_now,
                          'price_before': price_before,
                          'airline': airlines})

# open excel
workbook = openpyxl.load_workbook('C:\\Users\\huong.vu\\Desktop\\Git\\Google-Flights-Prices\\gmail_flight.xlsx')
# getting sheet in excel
sheet = workbook['Sheet1']
# getting last row in excel sheet
last_row = sheet.max_row

# getting historical data
hist_data = pd.DataFrame(data={'time': [sheet.cell(row=d, column=1).value for d in range(2, last_row + 1)],
                               'price_now': [sheet.cell(row=d, column=2).value for d in range(2, last_row + 1)],
                               'price_before': [sheet.cell(row=d, column=3).value for d in range(2, last_row + 1)],
                               'airline': [sheet.cell(row=d, column=4).value for d in range(2, last_row + 1)]})

# remove duplicates
insert_data = data.join(hist_data, lsuffix='_new', rsuffix='_hist')
insert_data = insert_data[pd.isnull(insert_data['time_hist'])].iloc[:, 0:4]

# write data into excel file
for k in range(last_row + 1, len(insert_data) + last_row + 1):
    try:
        sheet.cell(row=k, column=1).value = insert_data['time_new'].iloc[k - last_row - 1]
        sheet.cell(row=k, column=2).value = insert_data['price_now_new'].iloc[k - last_row - 1]
        sheet.cell(row=k, column=3).value = insert_data['price_before_new'].iloc[k - last_row - 1]
        sheet.cell(row=k, column=4).value = insert_data['airline_new'].iloc[k - last_row - 1]
    except Exception as e:
        print(str(e))

workbook.save('C:\\Users\\huong.vu\\Desktop\\Git\\Google-Flights-Prices\\gmail_flight.xlsx')
workbook.close()
